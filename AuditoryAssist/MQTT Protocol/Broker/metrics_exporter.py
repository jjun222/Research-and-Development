import csv
import os
import statistics
import sys
from collections import defaultdict

print("[metrics_exporter] python =", sys.executable)

try:
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ModuleNotFoundError as e:
    Workbook = None
    OPENPYXL_AVAILABLE = False
    print("⚠️ openpyxl import failed:", e)
    print("[metrics_exporter] sys.executable =", sys.executable)
    print("[metrics_exporter] sys.path =", sys.path)

BASE_DIR = "metrics"
TRIALS_CSV = os.path.join(BASE_DIR, "trial_results.csv")
XLSX_PATH = os.path.join(BASE_DIR, "metrics_summary.xlsx")

def _read_csv(path: str) -> list[dict]:
    if not os.path.exists(path):
        return []
    with open(path, "r", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))

def _to_float(x) -> float:
    try:
        return float(x)
    except Exception:
        return 0.0

def _to_int(x) -> int:
    try:
        return int(float(x))
    except Exception:
        return 0

def _pct(num: int, den: int) -> float:
    return round((num / den) * 100, 2) if den else 0.0

def _mean(vals) -> float:
    vals = [_to_float(v) for v in vals if str(v).strip() != ""]
    return round(sum(vals) / len(vals), 2) if vals else 0.0

def _p95(vals) -> float:
    vals = sorted([_to_float(v) for v in vals if str(v).strip() != ""])
    if not vals:
        return 0.0
    idx = max(0, min(len(vals) - 1, int(len(vals) * 0.95) - 1))
    return round(vals[idx], 2)

def _std(vals) -> float:
    vals = [_to_float(v) for v in vals if str(v).strip() != ""]
    if len(vals) < 2:
        return 0.0
    return round(statistics.stdev(vals), 2)

def _write_sheet(ws, rows: list[dict]) -> None:
    if not rows:
        ws.append(["NO DATA"])
        return
    header = list(rows[0].keys())
    ws.append(header)
    for row in rows:
        ws.append([row.get(h, "") for h in header])

def _write_csv_summary(path: str, rows: list[dict]) -> None:
    os.makedirs(BASE_DIR, exist_ok=True)
    if not rows:
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["NO DATA"])
        return

    header = list(rows[0].keys())
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=header)
        writer.writeheader()
        writer.writerows(rows)

def _summarize(rows: list[dict], group_name: str) -> list[dict]:
    grouped = defaultdict(list)
    for r in rows:
        if r.get("test_group") == group_name:
            grouped[(r.get("test_id"), r.get("scenario_id"))].append(r)

    out = []

    for (test_id, scenario_id), rs in grouped.items():
        repeats = len(rs)

        if group_name == "SR":
            success_count = sum(1 for r in rs if str(r.get("delivery_ok")).lower() == "true")
            failure_count = repeats - success_count
            lat = [r.get("e2e_latency_ms") for r in rs]
            out.append({
                "test_id": test_id,
                "scenario_id": scenario_id,
                "repeats": repeats,
                "success_count": success_count,
                "failure_count": failure_count,
                "delivery_success_rate_pct": _pct(success_count, repeats),
                "message_loss_rate_pct": _mean([r.get("message_loss_rate") for r in rs]),
                "duplicate_count_mean": _mean([r.get("duplicate_count") for r in rs]),
                "broker_processing_latency_mean_ms": _mean([r.get("broker_processing_latency_ms") for r in rs]),
                "e2e_latency_mean_ms": _mean(lat),
                "e2e_latency_p95_ms": _p95(lat),
                "jitter_ms": _std(lat),
                "throughput_mean_msg_s": _mean([r.get("throughput_msg_s") for r in rs]),
                "judgement": "PASS" if _pct(success_count, repeats) >= 95 else "CHECK",
                "notes": ""
            })

        elif group_name == "FP":
            success_count = sum(
                1 for r in rs
                if str(r.get("delivery_ok")).lower() == "true"
                and str(r.get("activation_ok")).lower() == "true"
                and str(r.get("ack_ok")).lower() == "true"
            )
            failure_count = repeats - success_count
            ack_success = sum(1 for r in rs if str(r.get("ack_ok")).lower() == "true")
            lat = [r.get("e2e_latency_ms") for r in rs]
            out.append({
                "test_id": test_id,
                "scenario_id": scenario_id,
                "repeats": repeats,
                "success_count": success_count,
                "failure_count": failure_count,
                "full_pass_success_rate_pct": _pct(success_count, repeats),
                "e2e_latency_mean_ms": _mean(lat),
                "e2e_latency_p95_ms": _p95(lat),
                "wrong_activation_count_total": sum(_to_int(r.get("wrong_activation_count")) for r in rs),
                "ack_success_rate_pct": _pct(ack_success, repeats),
                "judgement": "PASS" if _pct(success_count, repeats) >= 95 else "CHECK",
                "notes": ""
            })

        elif group_name == "PR":
            correct = sum(1 for r in rs if str(r.get("priority_ok")).lower() == "true")
            incorrect = repeats - correct
            out.append({
                "test_id": test_id,
                "scenario_id": scenario_id,
                "repeats": repeats,
                "correct_priority_count": correct,
                "incorrect_priority_count": incorrect,
                "priority_resolution_correctness_pct": _pct(correct, repeats),
                "priority_settle_time_mean_ms": _mean([r.get("priority_settle_time_ms") for r in rs]),
                "e2e_latency_mean_ms": _mean([r.get("e2e_latency_ms") for r in rs]),
                "emergency_dominance_success_rate_pct": _pct(correct, repeats),
                "judgement": "PASS" if _pct(correct, repeats) >= 95 else "CHECK",
                "notes": ""
            })

        elif group_name == "CM":
            receive_success = sum(1 for r in rs if str(r.get("delivery_ok")).lower() == "true")
            trigger_success = sum(
                1 for r in rs
                if str(r.get("all_true_ok")).lower() == "true"
                or str(r.get("activation_ok")).lower() == "true"
            )
            completeness = []
            for r in rs:
                expected_devices = _to_int(r.get("expected_devices"))
                activated_devices = _to_int(r.get("activated_devices"))
                completeness.append((activated_devices / expected_devices * 100) if expected_devices > 0 else 0)

            out.append({
                "test_id": test_id,
                "scenario_id": scenario_id,
                "repeats": repeats,
                "receive_success_rate_pct": _pct(receive_success, repeats),
                "message_loss_rate_pct": _mean([r.get("message_loss_rate") for r in rs]),
                "trigger_success_rate_pct": _pct(trigger_success, repeats),
                "activation_completeness_pct": _mean(completeness),
                "trigger_latency_mean_ms": _mean([r.get("trigger_latency_ms") for r in rs]),
                "e2e_latency_mean_ms": _mean([r.get("e2e_latency_ms") for r in rs]),
                "judgement": "PASS" if _pct(receive_success, repeats) >= 95 else "CHECK",
                "notes": ""
            })

    return out

def export_all() -> None:
    rows = _read_csv(TRIALS_CSV)

    sr = _summarize(rows, "SR")
    fp = _summarize(rows, "FP")
    pr = _summarize(rows, "PR")
    cm = _summarize(rows, "CM")

    # CSV summary는 항상 생성
    _write_csv_summary(os.path.join(BASE_DIR, "sr_summary.csv"), sr)
    _write_csv_summary(os.path.join(BASE_DIR, "fp_summary.csv"), fp)
    _write_csv_summary(os.path.join(BASE_DIR, "pr_summary.csv"), pr)
    _write_csv_summary(os.path.join(BASE_DIR, "cm_summary.csv"), cm)

    if not OPENPYXL_AVAILABLE:
        print("⚠️ openpyxl 없음: XLSX 대신 CSV summary만 생성")
        return

    os.makedirs(BASE_DIR, exist_ok=True)
    wb = Workbook()

    ws = wb.active
    ws.title = "SR_Summary"
    _write_sheet(ws, sr)

    ws = wb.create_sheet("FP_Summary")
    _write_sheet(ws, fp)

    ws = wb.create_sheet("PR_Summary")
    _write_sheet(ws, pr)

    ws = wb.create_sheet("CM_Summary")
    _write_sheet(ws, cm)

    ws = wb.create_sheet("Trial_Raw")
    _write_sheet(ws, rows)

    wb.save(XLSX_PATH)
    print("✅ metrics_summary.xlsx 저장 완료:", XLSX_PATH)

if __name__ == "__main__":
    export_all()
